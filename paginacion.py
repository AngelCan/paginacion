import os
import random
from time import sleep
from openpyxl import Workbook as NuevoExcel
from openpyxl import load_workbook as LeerExcel

class Paginacion:
    def __init__(self, tamMf, tamPag):
        self.memoFisica = tamMf
        self.tamPag = tamPag 
        self.numPags = int(tamMf/tamPag) 
        self.marcos = {} 
        self.memoriaFisica()
        self.memoriaLogica()
        self.tablaPags()

    def memoriaFisica(self):
        ws = NuevoExcel()
        wb = ws.active
        wb.cell(row=1, column=1, value="Marco")
        wb.cell(row=1, column=2, value="Direccion fisica")
        wb.cell(row=1, column=3, value="Proceso")

        for i in range(self.memoFisica):
            wb.cell(row=i+2, column=2, value=i)
        
        
        memo = random.randint(1, 2)

        marcos1 = [16, 4, 0, 12, 8]
        marcos2 = [18, 6, 2, 14, 21]
        marcos3 = [17, 5, 1, 11, 15]

        if (memo == 1):
            i = 0
            for f in marcos1:
                wb.cell(row=f+2, column=1, value=i)
                self.marcos.__setitem__(i, f)
                i += 1

        if (memo == 2):
            i = 0
            for f in marcos2:
                wb.cell(row=f+2, column=1, value=i)
                self.marcos.__setitem__(i, f)
                i += 1

        if (memo == 3):
            i = 0
            for f in marcos3:
                wb.cell(row=f+2, column=1, value=i)
                self.marcos.__setitem__(i, f)
                i += 1
        ws.save(os.path.dirname(__file__) + "/fisica.xlsx")

    def memoriaLogica(self):
        ws = NuevoExcel()
        wb = ws.active
        wb.cell(row=1, column=1, value="pagina")
        wb.cell(row=1, column=2, value="Direccion logica")
        wb.cell(row=1, column=3, value="Proceso")

        memoria_usada = 2*self.tamPag
        j = 0
        for i in range(self.memoFisica - memoria_usada):
            if (i % self.tamPag == 0):
                wb.cell(row=i+2, column=1, value=j)
                j += 1
            wb.cell(row=i+2, column=2, value=i)

        ws.save(os.path.dirname(__file__) + "/logica.xlsx")

    def tablaPags(self):
        ws = NuevoExcel()
        wb = ws.active
        wb.cell(row=1, column=1, value="pagina")
        wb.cell(row=1, column=2, value="marco")

        print("Tabla de paginas")
        print("Paginas/marcos")
        for i in self.marcos:
            print(i, self.marcos[i])
            wb.cell(row=i+2, column=1, value=i)
            wb.cell(row=i+2, column=2, value=self.marcos[i])
        ws.save(os.path.dirname(__file__) + "/tabla_paginas.xlsx")

    def guardarProc(self, Procesos):
        direc_fisica = self.marcos[Procesos.pagina]
        print("Direcion fisica: ", direc_fisica)

        lm = LeerExcel(os.path.dirname(__file__) + "/logica.xlsx") 
        logical = lm.active 
        logical['C' + str(Procesos.direc_logica + 2)] = Procesos.nombre
        lm.save(os.path.dirname(__file__) + "/logica.xlsx")
        
        pm = LeerExcel(os.path.dirname(__file__) + "/fisica.xlsx")
        physical = pm.active
        for i in range(self.tamPag):
            physical['C' + str(direc_fisica + 2 + i)] = Procesos.nombre
        pm.save(os.path.dirname(__file__) + "/fisica.xlsx")

class Procesos:
    def __init__(self, nombre, tamMf, tamPag):
        self.nombre = nombre
        self.direc_logica = random.randint(0, (tamMf - 1) - tamPag*4) 
        self.pagina = int(self.direc_logica/tamPag) 
        self.desplazamiento = self.direc_logica - self.pagina*tamPag
        print("\nProceso: ", self.nombre)
        print("Pag.: " +str(self.pagina) + " |Direccion. Logica.: " +str(self.direc_logica)+ " |Compensacion: " +str(self.desplazamiento))
        sleep(1)

def crearProcesos(tamMf, tamPag, pm):
    print("\n\nCreando procesos...")
    # Creamos procesos aleatorios
    for i in range(6):
        p1 = Procesos("Proc. " + str(i), tamMf, tamPag)
        try:
            pm.guardarProc(p1)
        except:
            pass
        
if __name__ == '__main__':
    tamMf = 32
    tamPag = 4
    print("Memoria: " + str(tamMf) + " bytes")
    print("Tamanio de pagina: " + str(tamPag) + " bytes\n")

    p = Paginacion(tamMf, tamPag)
    crearProcesos(tamMf, tamPag, p)